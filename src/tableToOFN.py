from typing import List
import openpyxl  # type: ignore
import sys
from ofnClasses import Relationship, Term, TermClass, Trope, Vocabulary, VocabularyType, getClass, getTrope, ClassType
from outputToRDF import convertToRDF
from ofnBindings import *
import csv
import warnings


def openXLSX(file: str):
    return openpyxl.load_workbook(file, data_only=True)


def csvToSheets(vcCSV, soCSV, itCSV, rlCSV):
    wb = openpyxl.Workbook()
    vcSheet = wb.create_sheet(SHEET_VOCABULARY)
    soSheet = wb.create_sheet(SHEET_CLASS)
    itSheet = wb.create_sheet(SHEET_TROPE)
    rlSheet = wb.create_sheet(SHEET_RELATIONSHIP)

    for x in [(vcCSV, vcSheet), (soCSV, soSheet), (itCSV, itSheet), (rlCSV, rlSheet)]:
        if not x[0].endswith("csv"):
            raise Exception()
        with open(x[0]) as f:
            reader = csv.reader(f, delimiter=",")
            for row in reader:
                x[1].append(row)

    return (soSheet, itSheet, rlSheet, vcSheet)


def xlsxToSheets(file: str):
    soSheet = None
    itSheet = None
    rlSheet = None
    vcSheet = None
    if file.endswith("xlsx"):
        wb = openpyxl.load_workbook(file, data_only=True)
        vcSheet = wb[SHEET_VOCABULARY]
        soSheet = wb[SHEET_CLASS]
        itSheet = wb[SHEET_TROPE]
        rlSheet = wb[SHEET_RELATIONSHIP]

    if soSheet is not None and itSheet is not None and rlSheet is not None and vcSheet is not None:
        return (soSheet, itSheet, rlSheet, vcSheet)
    else:
        raise Exception()


def soSheetToOFN(sheet) -> List[TermClass]:
    header = True
    nameIndex = -1
    descriptionIndex = -1
    definitionIndex = -1
    sourceIndex = -1
    subClassOfIndex = -1
    equivalentIndex = -1
    aisIndex = -1
    agendaIndex = -1
    iriIndex = -1
    typeIndex = -1
    sos = []
    for lst in sheet:
        row = [cell.value for cell in lst]
        if header:
            nameIndex = row.index(OFN_NAME)
            descriptionIndex = row.index(OFN_DESCRIPTION)
            definitionIndex = row.index(OFN_DEFINITION)
            sourceIndex = row.index(OFN_SOURCE)
            subClassOfIndex = row.index(OFN_SUBCLASS)
            equivalentIndex = row.index(OFN_EQUIVALENT)
            iriIndex = row.index(OFN_IRI)
            aisIndex = row.index(OFN_RPP_AIS)
            agendaIndex = row.index(OFN_RPP_AGENDA)
            typeIndex = row.index(OFN_TYPE)
            relatedSourceIndex = row.index(OFN_RELATED)
            alternativeNameIndex = row.index(OFN_ALTERNATIVE)
            header = False
        else:
            term = TermClass()
            if row[nameIndex] is None:
                warnings.warn("warn")
                continue
            term.name = {DEFAULT_LANGUAGE: row[nameIndex]}
            if (row[typeIndex]):
                if row[typeIndex].strip().lower() == OFN_SUBJECT.lower():
                    term.type = ClassType.SUBJECT
                elif row[typeIndex].strip().lower() == OFN_OBJECT.lower():
                    term.type = ClassType.OBJECT
                else:
                    warnings.warn("warn")
            if row[definitionIndex]:
                term.definition = {DEFAULT_LANGUAGE: row[definitionIndex]}
            if row[descriptionIndex]:
                term.description = {DEFAULT_LANGUAGE: row[descriptionIndex]}
            if row[sourceIndex]:
                term.source = row[sourceIndex]
            if row[subClassOfIndex]:
                term.subClassOf.append(row[subClassOfIndex])
            if row[equivalentIndex]:
                term.equivalent.append(row[equivalentIndex])
            if row[iriIndex]:
                term.iri = row[iriIndex]
            if row[aisIndex]:
                term.ais = row[aisIndex]
            if row[agendaIndex]:
                term.agenda = row[agendaIndex]
            if row[relatedSourceIndex]:
                term.related += [x.strip()
                                 for x in row[relatedSourceIndex].split(MULTIPLE_VALUE_SEPARATOR)]
            if row[alternativeNameIndex]:
                term.alternateName += [(DEFAULT_LANGUAGE, x.strip())
                                       for x in row[alternativeNameIndex].split(MULTIPLE_VALUE_SEPARATOR)]
            sos.append(term)
    return sos


def itSheetToOFN(sheet) -> List[Trope]:
    header = True
    termClassIndex = -1
    datatypeIndex = -1
    nameIndex = -1
    descriptionIndex = -1
    definitionIndex = -1
    sourceIndex = -1
    subClassOfIndex = -1
    equivalentIndex = -1
    sharedInPPDFIndex = -1
    rppTypeIndex = -1
    rppPrivateTypeSourceIndex = -1
    iriIndex = -1
    tropes = []
    for lst in sheet:
        row = [cell.value for cell in lst]
        if header:
            nameIndex = row.index(OFN_NAME)
            descriptionIndex = row.index(OFN_DESCRIPTION)
            definitionIndex = row.index(OFN_DEFINITION)
            sourceIndex = row.index(OFN_SOURCE)
            subClassOfIndex = row.index(OFN_SUBCLASS)
            equivalentIndex = row.index(OFN_EQUIVALENT)
            iriIndex = row.index(OFN_IRI)
            termClassIndex = row.index(OFN_SUBJECT_OR_OBJECT)
            datatypeIndex = row.index(OFN_DATATYPE)
            sharedInPPDFIndex = row.index(OFN_RPP_SHARED)
            rppTypeIndex = row.index(OFN_RPP_TYPE)
            rppPrivateTypeSourceIndex = row.index(OFN_RPP_PRIVATE_SOURCE)
            relatedSourceIndex = row.index(OFN_RELATED)
            alternativeNameIndex = row.index(OFN_ALTERNATIVE)
            header = False
        else:
            if row[nameIndex] is None or row[termClassIndex] is None:
                continue
            term = Trope()
            term.name = {DEFAULT_LANGUAGE: row[nameIndex]}
            term.target = row[termClassIndex]
            if row[definitionIndex]:
                term.definition = {DEFAULT_LANGUAGE: row[definitionIndex]}
            if row[descriptionIndex]:
                term.description = {DEFAULT_LANGUAGE: row[descriptionIndex]}
            if row[sourceIndex]:
                term.source = row[sourceIndex]
            if row[subClassOfIndex]:
                term.subClassOf.append(row[subClassOfIndex])
            if row[equivalentIndex]:
                term.equivalent.append(row[equivalentIndex])
            if row[iriIndex]:
                term.iri = row[iriIndex]
            if row[sharedInPPDFIndex]:
                term.sharedInPPDF = row[sharedInPPDFIndex]
            if row[datatypeIndex]:
                term.datatype = row[datatypeIndex]
            if row[rppTypeIndex]:
                term.rppType = row[rppTypeIndex]
            if row[rppPrivateTypeSourceIndex]:
                term.rppPrivateTypeSource = row[rppPrivateTypeSourceIndex]
            if row[relatedSourceIndex]:
                term.related += [x.strip()
                                 for x in row[relatedSourceIndex].split(MULTIPLE_VALUE_SEPARATOR)]
            if row[alternativeNameIndex]:
                term.alternateName += [(DEFAULT_LANGUAGE, x.strip())
                                       for x in row[alternativeNameIndex].split(MULTIPLE_VALUE_SEPARATOR)]
            tropes.append(term)
    return tropes


def rlSheetToOFN(sheet) -> List[Relationship]:
    header = True
    termClassSourceIndex = -1
    termClassTargetIndex = -1
    nameIndex = -1
    descriptionIndex = -1
    definitionIndex = -1
    sourceIndex = -1
    subClassOfIndex = -1
    equivalentIndex = -1
    iriIndex = -1
    relationships = []
    for lst in sheet:
        row = [cell.value for cell in lst]
        if header:
            nameIndex = row.index(OFN_NAME)
            descriptionIndex = row.index(OFN_DESCRIPTION)
            definitionIndex = row.index(OFN_DEFINITION)
            sourceIndex = row.index(OFN_SOURCE)
            subClassOfIndex = row.index(OFN_SUBCLASS)
            equivalentIndex = row.index(OFN_EQUIVALENT)
            iriIndex = row.index(OFN_IRI)
            termClassIndices = [x for x, y in enumerate(
                row) if y == OFN_SUBJECT_OR_OBJECT]
            termClassSourceIndex = termClassIndices[0]
            termClassTargetIndex = termClassIndices[1]
            relatedSourceIndex = row.index(OFN_RELATED)
            alternativeNameIndex = row.index(OFN_ALTERNATIVE)
            sharedInPPDFIndex = row.index(OFN_RPP_SHARED)
            rppTypeIndex = row.index(OFN_RPP_TYPE)
            rppPrivateTypeSourceIndex = row.index(OFN_RPP_PRIVATE_SOURCE)

            header = False
        else:
            if row[nameIndex] is None or row[termClassSourceIndex] is None or row[termClassTargetIndex] is None:
                continue
            term = Relationship(row[termClassSourceIndex],
                                row[termClassTargetIndex])
            term.name = {DEFAULT_LANGUAGE: row[nameIndex]}
            if row[definitionIndex]:
                term.definition = {DEFAULT_LANGUAGE: row[definitionIndex]}
            if row[descriptionIndex]:
                term.description = {DEFAULT_LANGUAGE: row[descriptionIndex]}
            if row[sourceIndex]:
                term.source = row[sourceIndex]
            if row[subClassOfIndex]:
                term.subClassOf.append(row[subClassOfIndex])
            if row[equivalentIndex]:
                term.equivalent.append(row[equivalentIndex])
            if row[iriIndex]:
                term.iri = row[iriIndex]
            if row[sharedInPPDFIndex]:
                if row[sharedInPPDFIndex].strip().lower() == YES.lower():
                    term.sharedInPPDF = True
                elif row[sharedInPPDFIndex].strip().lower() == NO.lower():
                    term.sharedInPPDF = False
                else:
                    warnings.warn("warn")
            if row[rppTypeIndex]:
                term.rppType = row[rppTypeIndex]
            if row[rppPrivateTypeSourceIndex]:
                term.rppPrivateTypeSource = row[rppPrivateTypeSourceIndex]
            if row[relatedSourceIndex]:
                term.related += [x.strip()
                                 for x in row[relatedSourceIndex].split(MULTIPLE_VALUE_SEPARATOR)]
            if row[alternativeNameIndex]:
                term.alternateName += [(DEFAULT_LANGUAGE, x.strip())
                                       for x in row[alternativeNameIndex].split(MULTIPLE_VALUE_SEPARATOR)]
            relationships.append(term)
    return relationships


def vcSheetToOFN(sheet):
    name = None
    desc = None
    lkod = None
    for lst in sheet:
        row = [cell.value for cell in lst]
        if name is None:
            name = row[1]
            continue
        elif desc is None:
            desc = row[1]
            continue
        elif lkod is None:
            lkod = row[1]
            continue
        else:
            break
    if name is None or desc is None or lkod is None:
        raise Exception()
    return (name, desc, lkod)


inputLocation = sys.argv[1]
outputLocation = sys.argv[2]


def tableToOFN():
    soSheet = None
    itSheet = None
    rlSheet = None
    vcSheet = None
    if len(sys.argv) == 3:
        (soSheet, itSheet, rlSheet, vcSheet) = xlsxToSheets(inputLocation)
    elif len(sys.argv) == 6:
        (soSheet, itSheet, rlSheet, vcSheet) = csvToSheets(inputLocation)
    else:
        raise Exception()
    soList = soSheetToOFN(soSheet)
    itList = itSheetToOFN(itSheet)
    rlList = rlSheetToOFN(rlSheet)
    (name, desc, lkod) = vcSheetToOFN(vcSheet)
    vocabulary = Vocabulary()
    vocabulary.name[DEFAULT_LANGUAGE] = name
    vocabulary.description[DEFAULT_LANGUAGE] = desc
    vocabulary.lkod = lkod
    vocabulary.terms.extend(soList)
    vocabulary.terms.extend(itList)
    vocabulary.terms.extend(rlList)
    if len(itList) > 0 or len(rlList) > 0:
        vocabulary.type = VocabularyType.CONCEPTUAL_MODEL
    vocabulary.name = {DEFAULT_LANGUAGE: "test"}
    convertToRDF(vocabulary, DEFAULT_LANGUAGE, outputLocation)


tableToOFN()
