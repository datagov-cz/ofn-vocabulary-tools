from typing import List
import openpyxl
import sys
from ofnClasses import Relationship, TermClass, Trope, Vocabulary, VocabularyType, getClass, getTrope, ClassType
from outputToRDF import convertToRDF

inputLocation = sys.argv[1]
outputLocation = sys.argv[2]
defaultLanguage = "cs"

# TODO: CSV support


def xlsxToSheets(file: str):
    soSheet = None
    itSheet = None
    rlSheet = None
    vcSheet = None
    if file.endswith("xlsx"):
        wb = openpyxl.load_workbook(file, data_only=True)
        vcSheet = wb["Slovník"]
        soSheet = wb["Subjekty a objekty práva"]
        itSheet = wb["Vlastnosti"]
        rlSheet = wb["Vztahy"]

    if soSheet is not None and itSheet is not None and rlSheet is not None and vcSheet is not None:
        return (soSheet, itSheet, rlSheet, vcSheet)
    else:
        raise Exception()


def soSheetToOFN(sheet) -> List[TermClass]:
    header = True
    nameIndex = None
    descriptionIndex = None
    definitionIndex = None
    sourceIndex = None
    subClassOfIndex = None
    equivalentIndex = None
    aisIndex = None
    agendaIndex = None
    iriIndex = None
    typeIndex = None
    list = []
    for lst in sheet:
        row = [cell.value for cell in lst]
        if header:
            nameIndex = row.index("Název")
            descriptionIndex = row.index("Popis")
            definitionIndex = row.index("Definice")
            sourceIndex = row.index("Zdroj")
            subClassOfIndex = row.index("Nadřazený pojem")
            equivalentIndex = row.index("Ekvivalentní pojem")
            iriIndex = row.index("Identifikátor")
            aisIndex = row.index("Agendový informační systém")
            agendaIndex = row.index("Agenda")
            typeIndex = row.index("Typ")
            header = False
        else:
            term = TermClass()
            if row[nameIndex] is None:
                continue
            term.name = {defaultLanguage: row[nameIndex]}
            if (row[typeIndex]):
                if row[typeIndex].strip().lower() == "subjekt práva":
                    term.type = ClassType.SUBJECT
                elif row[typeIndex].strip().lower() == "objekt práva":
                    term.type = ClassType.OBJECT
            if row[definitionIndex]:
                term.definition = {defaultLanguage: row[definitionIndex]}
            if row[descriptionIndex]:
                term.description = {defaultLanguage: row[descriptionIndex]}
            if row[sourceIndex]:
                term.source = row[sourceIndex]
            if row[subClassOfIndex]:
                term.subClassOf.append(row[subClassOfIndex])
            if row[equivalentIndex]:
                term.equivalent.append(row[equivalentIndex])
            if row[iriIndex]:
                term.iri = row[iriIndex]
            if row[aisIndex]:
                term.ais = row[aisIndex]
            if row[agendaIndex]:
                term.agenda = row[agendaIndex]
            list.append(term)
    return list


def itSheetToOFN(sheet) -> List[Trope]:
    header = True
    termClassIndex = None
    datatypeIndex = None
    nameIndex = None
    descriptionIndex = None
    definitionIndex = None
    sourceIndex = None
    subClassOfIndex = None
    equivalentIndex = None
    sharedInPpdfIndex = None
    rppTypeIndex = None
    rppPrivateTypeSourceIndex = None
    iriIndex = None
    list = []
    for lst in sheet:
        row = [cell.value for cell in lst]
        if header:
            nameIndex = row.index("Název")
            descriptionIndex = row.index("Popis")
            definitionIndex = row.index("Definice")
            sourceIndex = row.index("Zdroj")
            subClassOfIndex = row.index("Nadřazený pojem")
            equivalentIndex = row.index("Ekvivalentní pojem")
            iriIndex = row.index("Identifikátor")
            termClassIndex = row.index("Subjekt nebo objekt práva")
            datatypeIndex = row.index("Datový typ")
            sharedInPpdfIndex = row.index("Je pojem sdílen v PPDF?")
            rppTypeIndex = row.index("Je pojem veřejný?")
            rppPrivateTypeSourceIndex = row.index(
                "Ustanovení dokládající neveřejnost pojmu")
            header = False
        else:
            if row[nameIndex] is None or row[termClassIndex] is None:
                continue
            term = Trope()
            term.name = {defaultLanguage: row[nameIndex]}
            term.target = row[termClassIndex]
            if row[definitionIndex]:
                term.definition = {defaultLanguage: row[definitionIndex]}
            if row[descriptionIndex]:
                term.description = {defaultLanguage: row[descriptionIndex]}
            if row[sourceIndex]:
                term.source = row[sourceIndex]
            if row[subClassOfIndex]:
                term.subClassOf.append(row[subClassOfIndex])
            if row[equivalentIndex]:
                term.equivalent.append(row[equivalentIndex])
            if row[iriIndex]:
                term.iri = row[iriIndex]
            if row[sharedInPpdfIndex]:
                term.sharedInPPDF = row[sharedInPpdfIndex]
            if row[datatypeIndex]:
                term.datatype = row[datatypeIndex]
            if row[rppTypeIndex]:
                term.rppType = row[rppTypeIndex]
            if row[rppPrivateTypeSourceIndex]:
                term.rppPrivateTypeSource = row[rppPrivateTypeSourceIndex]
            list.append(term)
    return list


def rlSheetToOFN(sheet) -> List[Relationship]:
    header = True
    termClassSourceIndex = None
    termClassTargetIndex = None
    nameIndex = None
    descriptionIndex = None
    definitionIndex = None
    sourceIndex = None
    subClassOfIndex = None
    equivalentIndex = None
    iriIndex = None
    list = []
    for lst in sheet:
        row = [cell.value for cell in lst]
        if header:
            nameIndex = row.index("Název")
            descriptionIndex = row.index("Popis")
            definitionIndex = row.index("Definice")
            sourceIndex = row.index("Zdroj")
            subClassOfIndex = row.index("Nadřazený pojem")
            equivalentIndex = row.index("Ekvivalentní pojem")
            iriIndex = row.index("Identifikátor")
            termClassIndices = [x for x, y in enumerate(
                row) if y == "Subjekt nebo objekt práva"]
            termClassSourceIndex = termClassIndices[0]
            termClassTargetIndex = termClassIndices[1]
            header = False
        else:
            if row[nameIndex] is None or row[termClassSourceIndex] is None or row[termClassTargetIndex] is None:
                continue
            term = Relationship(row[termClassSourceIndex],
                                row[termClassTargetIndex])
            term.name = {defaultLanguage: row[nameIndex]}
            if row[definitionIndex]:
                term.definition = {defaultLanguage: row[definitionIndex]}
            if row[descriptionIndex]:
                term.description = {defaultLanguage: row[descriptionIndex]}
            if row[sourceIndex]:
                term.source = row[sourceIndex]
            if row[subClassOfIndex]:
                term.subClassOf.append(row[subClassOfIndex])
            if row[equivalentIndex]:
                term.equivalent.append(row[equivalentIndex])
            if row[iriIndex]:
                term.iri = row[iriIndex]
            list.append(term)
    return list


def vcSheetToOFN(sheet):
    name = None
    desc = None
    lkod = None
    for lst in sheet:
        row = [cell.value for cell in lst]
        if name is None:
            name = row[1]
            continue
        elif desc is None:
            desc = row[1]
            continue
        elif lkod is None:
            lkod = row[1]
            continue
        else:
            break
    if name is None or desc is None or lkod is None:
        raise Exception()
    return (name, desc, lkod)


def tableToOFN():
    (soSheet, itSheet, rlSheet, vcSheet) = xlsxToSheets(inputLocation)
    soList = soSheetToOFN(soSheet)
    itList = itSheetToOFN(itSheet)
    rlList = rlSheetToOFN(rlSheet)
    (name, desc, lkod) = vcSheetToOFN(vcSheet)
    vocabulary = Vocabulary()
    vocabulary.name[defaultLanguage] = name
    vocabulary.description[defaultLanguage] = desc
    vocabulary.lkod = lkod
    vocabulary.terms.extend(soList)
    vocabulary.terms.extend(itList)
    vocabulary.terms.extend(rlList)
    if len(itList) > 0 or len(rlList) > 0:
        vocabulary.type = VocabularyType.CONCEPTUAL_MODEL

    vocabulary.name = {defaultLanguage: "test"}
    convertToRDF(vocabulary, defaultLanguage, outputLocation)


tableToOFN()

# def tableToOFN(soSheet, itSheet, rlSheet):

#     convertToRDF(vocabulary, defaultLanguage, outputLocation)

# def csvToSheets(soFile: str, itFile: str, rlFile: str):
#     soExtension = os.path.splitext(soFile)
# itExtension = os.path.splitext(itFile)
# rlExtension = os.path.splitext(rlFile)
# soSheet = None
# itSheet = None
# rlSheet = None
# if "csv" in soExtension[1]:
#     with open(soFile) as soCSV:
#         for row in soCSV:
#             print([cell for cell in row])
# if itExtension is "csv":
#     pass
# if rlExtension is "csv":
#     pass

# if soSheet is not None and itSheet is not None and rlSheet is not None:
#     return (soSheet, itSheet, rlSheet)
# else:
#     raise Exception()
# wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
# soSheet = wb["Subjekty a objekty práva"]
# itSheet = wb["Vlastnosti"]
# rlSheet = wb["Vztahy"]
# if (soSheet is not None):
#     for row in soSheet.iter_rows():
#         print([cell.value for cell in row])
# if (itSheet is not None):
#     pass
# if (rlSheet is not None):
#     pass
